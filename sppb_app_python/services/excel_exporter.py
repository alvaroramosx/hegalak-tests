from __future__ import annotations

from pathlib import Path
from typing import Dict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def export_to_excel(output_path: str, patient_info: Dict[str, str], results: Dict[str, object]) -> str:
	file_path = Path(output_path)
	file_path.parent.mkdir(parents=True, exist_ok=True)

	wb = Workbook()
	ws = wb.active
	ws.title = "SPPB"

	rows = [
		["Fecha", patient_info.get("date", "")],
		["Nombre", patient_info.get("name", "")],
		["Edad", patient_info.get("age", "")],
		[],
		["Equilibrio"],
		["Pies juntos (s)", results.get("feet_together_s", "")],
		["Semitándem (s)", results.get("semi_tandem_s", "")],
		["Tándem (s)", results.get("tandem_s", "")],
		["Puntuación equilibrio", results.get("balance_score", "")],
		[],
		["Marcha 4 m"],
		["Tiempo (s)", "No pudo" if results.get("gait_unable", False) else results.get("gait_time_s", "")],
		["Puntuación marcha", results.get("gait_score", "")],
		[],
		["Levantarse de la silla (5x)"],
		["Tiempo (s)", "No pudo" if results.get("chair_unable", False) else results.get("chair_time_s", "")],
		["Puntuación silla", results.get("chair_score", "")],
		[],
		["Total SPPB", results.get("total", "")],
		["Conclusión", results.get("interpretation", "")],
	]

	for r_idx, row in enumerate(rows, start=1):
		for c_idx, value in enumerate(row, start=1):
			ws.cell(row=r_idx, column=c_idx, value=value)

	# Aplicar fuente negra a todas las celdas usadas
	for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
		for cell in row:
			cell.font = Font(color="000000")

	for col in range(1, ws.max_column + 1):
		letter = get_column_letter(col)
		max_len = 0
		for row in range(1, ws.max_row + 1):
			v = ws.cell(row=row, column=col).value
			max_len = max(max_len, len(str(v)) if v is not None else 0)
		ws.column_dimensions[letter].width = min(60, max(12, max_len + 2))

	wb.save(str(file_path))
	return str(file_path.resolve())

